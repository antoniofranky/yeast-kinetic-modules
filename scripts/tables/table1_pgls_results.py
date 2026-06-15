"""Build PGLS phenotype correlations table (source data for Table 1 in manuscript).
Network properties are ln-transformed; biomass yield and substrate count
are on the linear scale. Signed Pearson r = sign(slope) * sqrt(R^2)."""
import csv
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

REPO_ROOT = Path(__file__).resolve().parents[2]
CSV_DIR   = str(REPO_ROOT / "results" / "pgls")
OUT       = str(REPO_ROOT / "supplementary" / "Table1_pgls.xlsx")

wb = Workbook()

# ---------------------------------------------------------------------------
# Sheet 1: Legend
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Legend"
ws["A1"] = "Table 1: PGLS phenotype correlations"
ws["A1"].font = Font(bold=True, size=13)
ws["A3"] = (
    "Phylogenetic Generalized Least Squares (PGLS) regression results for the "
    "relationship between metabolic network properties and two phenotypic "
    "traits (biomass yield, substrate usage breadth) across budding yeast "
    "species. Network properties (total reaction count, giant kinetic module "
    "size, ACR metabolite count) are ln-transformed; biomass yield and "
    "substrate count are on the linear scale. Pagel's lambda was estimated "
    "by maximum likelihood; phylogeny from Shen et al. (2018). Pearson r is "
    "signed: r = sign(slope) * sqrt(R^2)."
)
ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[3].height = 80
ws.column_dimensions["A"].width = 110

ws["A5"] = "Sheets:"
ws["A5"].font = Font(bold=True)
ws["A6"] = "Direct_PGLS  - univariate PGLS for each (predictor, response) pair"
ws["A7"] = "Joint_PGLS   - multiple-predictor PGLS testing partial effects, with AIC comparison and likelihood-ratio tests"

ws["A9"] = "Datasets:"
ws["A9"].font = Font(bold=True)
ws["A10"] = "Full   - all 316 species with complete measurements"
ws["A11"] = "Clean  - 312 species (4 outliers with <1% reactions in giant module excluded)"
ws["A12"] = "All joint models use the Clean dataset (n = 312)."

ws["A14"] = "Significance: *** p<0.001, ** p<0.01, * p<0.05, . p<0.1"
ws["A14"].font = Font(italic=True)

# ---------------------------------------------------------------------------
# Sheet 2: Direct (univariate) PGLS
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Direct_PGLS")

header = ["Response", "Predictor", "Dataset", "n", "slope",
          "Pearson_r", "R_squared", "lambda", "p_value", "Significance"]
ws.append(header)

def sig_code(p):
    if p is None or p == "":
        return ""
    p = float(p)
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    if p < 0.1:   return "."
    return ""

def parse_comparison(s):
    """Split 'Y ~ X' into (response, predictor)."""
    parts = [p.strip() for p in s.split("~")]
    return parts[0], parts[1] if len(parts) > 1 else ""

with open(os.path.join(CSV_DIR, "pgls_results_kinetic_phenotype.csv"), encoding="utf-8") as fh:
    reader = csv.DictReader(fh)
    rows = list(reader)

def sort_key(r):
    is_biomass = "Biomass yield" in r["comparison"]
    return (0 if is_biomass else 1, r["comparison"], r["dataset"])

rows.sort(key=sort_key)

for r in rows:
    response, predictor = parse_comparison(r["comparison"])
    p = float(r["p_value"])
    ws.append([
        response,
        predictor,
        r["dataset"],
        int(r["n"]),
        float(r["slope"]),
        float(r["pearson_r"]),
        float(r["r_squared"]),
        float(r["lambda"]),
        p,
        sig_code(p),
    ])

# Format
for col_letter, width in zip(
        ["A","B","C","D","E","F","G","H","I","J"],
        [24,34,9,7,14,12,12,10,12,14]):
    ws.column_dimensions[col_letter].width = width
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="E0E0E0")

# Scientific format for slope, R^2, p_value; fixed decimals for r and lambda
for row in ws.iter_rows(min_row=2):
    row[4].number_format = "0.00E+00"   # slope
    row[5].number_format = "0.000"      # Pearson r
    row[6].number_format = "0.0000"     # R^2
    row[7].number_format = "0.000"      # lambda
    row[8].number_format = "0.00E+00"   # p_value

# ---------------------------------------------------------------------------
# Sheet 3: Joint PGLS
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Joint_PGLS")

def write_section(start_row, title, predictors_csv, aic_csv, lrt_chi2, lrt_p):
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=11)
    r = start_row + 1
    ws.cell(row=r, column=1, value="Joint model coefficients").font = Font(italic=True)
    r += 1
    headers = ["predictor", "estimate", "std_error", "t_value", "p_value", "significance", "VIF"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E0E0E0")
    r += 1

    with open(os.path.join(CSV_DIR, predictors_csv), encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            p = float(row["p_value"]) if row["p_value"] else None
            vif = row.get("vif", "NA")
            try: vif = float(vif) if vif not in ("NA", "") else None
            except: vif = None
            ws.cell(row=r, column=1, value=row["predictor"])
            ws.cell(row=r, column=2, value=float(row["estimate"])).number_format = "0.000E+00"
            ws.cell(row=r, column=3, value=float(row["std_error"])).number_format = "0.000E+00"
            ws.cell(row=r, column=4, value=float(row["t_value"])).number_format = "0.000"
            ws.cell(row=r, column=5, value=p).number_format = "0.000E+00"
            ws.cell(row=r, column=6, value=sig_code(p) if p is not None else "")
            if vif is not None:
                ws.cell(row=r, column=7, value=vif).number_format = "0.00"
            r += 1

    r += 1
    if aic_csv:
        ws.cell(row=r, column=1, value="AIC comparison").font = Font(italic=True)
        r += 1
        aic_headers = ["model", "n_params", "AIC", "R_squared", "delta_AIC"]
        for ci, h in enumerate(aic_headers, start=1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="F0F0F0")
        r += 1
        with open(os.path.join(CSV_DIR, aic_csv), encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                ws.cell(row=r, column=1, value=row["model"])
                ws.cell(row=r, column=2, value=int(row["n_params"]))
                ws.cell(row=r, column=3, value=float(row["AIC"])).number_format = "0.00"
                ws.cell(row=r, column=4, value=float(row["R2"])).number_format = "0.000"
                ws.cell(row=r, column=5, value=float(row["delta_AIC"])).number_format = "0.00"
                r += 1
        r += 1

    if lrt_chi2 is not None:
        ws.cell(row=r, column=1, value="Likelihood-ratio test (joint vs size-only)").font = Font(italic=True)
        r += 1
        ws.cell(row=r, column=1, value="chi_squared (df=1)")
        ws.cell(row=r, column=2, value=lrt_chi2).number_format = "0.000"
        r += 1
        ws.cell(row=r, column=1, value="p_value")
        ws.cell(row=r, column=2, value=lrt_p).number_format = "0.0000"
        r += 1
    return r + 2

r = 1
r = write_section(
    r,
    "Model 1: biomass_yield ~ ln(n_total_split_reactions) + ln(n_giant_reactions)",
    "pgls_joint_biomass.csv",
    "pgls_joint_aic_comparison.csv",
    lrt_chi2=3.398,
    lrt_p=0.0653,
)
r = write_section(
    r,
    "Model 2: biomass_yield ~ ln(n_total_split_reactions) + ln(n_acr)",
    "pgls_joint_biomass_acr.csv",
    aic_csv=None,
    lrt_chi2=0.578,
    lrt_p=0.4471,
)
r = write_section(
    r,
    "Model 3: n_substrates ~ ln(n_total_split_reactions) + ln(n_giant_reactions)",
    "pgls_joint_substrate.csv",
    aic_csv=None,
    lrt_chi2=1.043,
    lrt_p=0.3070,
)

ws.cell(row=r, column=1, value=(
    "Notes: All models use the Clean dataset (n = 312, 4 outlier species excluded). "
    "Network counts are ln-transformed. VIF = variance inflation factor (predictor "
    "collinearity); VIF > 5 indicates partial coefficients may be unstable, in which "
    "case the LRT and AIC comparison provide more robust evidence than individual "
    "partial p-values. log_total_split_reactions = ln of elementary-reaction-step "
    "count; log_giant_reactions = ln of elementary steps in the giant kinetic "
    "module; log_acr = ln of ACR metabolite count; n_substrates = substrate usage "
    "breadth (Biolog assay, linear scale)."
)).alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.row_dimensions[r].height = 110

for col_letter, width in zip(["A","B","C","D","E","F","G"], [44,14,14,12,14,14,10]):
    ws.column_dimensions[col_letter].width = width

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Sheets: {wb.sheetnames}")
